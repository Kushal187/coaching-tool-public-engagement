import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

DARK_BLUE = "1B3A5C"
MED_BLUE = "2E5E8E"
LIGHT_BLUE = "D6E4F0"
ACCENT_GREEN = "C6EFCE"
ACCENT_YELLOW = "FFF2CC"
ACCENT_RED = "FFC7CE"
ACCENT_ORANGE = "FCE4D6"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"

title_font = Font(name="Calibri", size=14, bold=True, color=WHITE)
subtitle_font = Font(name="Calibri", size=11, italic=True, color=WHITE)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
body_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", size=10, bold=True)
small_font = Font(name="Calibri", size=9)

title_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
header_fill = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
green_fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid")
yellow_fill = PatternFill(start_color=ACCENT_YELLOW, end_color=ACCENT_YELLOW, fill_type="solid")
red_fill = PatternFill(start_color=ACCENT_RED, end_color=ACCENT_RED, fill_type="solid")
orange_fill = PatternFill(start_color=ACCENT_ORANGE, end_color=ACCENT_ORANGE, fill_type="solid")
gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

ACCENT_PURPLE = "E2D0F8"
purple_fill = PatternFill(start_color=ACCENT_PURPLE, end_color=ACCENT_PURPLE, fill_type="solid")

STATUS_FILLS = {
    "Completed": green_fill,
    "Started": yellow_fill,
    "Not Started": red_fill,
    "Confirmed": green_fill,
    "Planned": yellow_fill,
    "Starting here": orange_fill,
    "To evaluate": orange_fill,
    "Submitted": green_fill,
    "Pending": orange_fill,
    "Scheduled": yellow_fill,
    "Pivoted": purple_fill,
}


def apply_title_row(ws, row, text, cols=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(vertical="center")
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).fill = title_fill


def apply_subtitle_row(ws, row, text, cols=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = subtitle_font
    cell.fill = title_fill
    cell.alignment = Alignment(vertical="center")
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).fill = title_fill


def apply_headers(ws, row, headers):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def apply_data_row(ws, row, values, is_alt=False):
    fill = gray_fill if is_alt else None
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = body_font
        cell.alignment = wrap_align
        cell.border = thin_border
        if fill:
            cell.fill = fill


def apply_status_color(ws, row, col):
    cell = ws.cell(row=row, column=col)
    val = str(cell.value or "").strip()
    if val in STATUS_FILLS:
        cell.fill = STATUS_FILLS[val]
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.alignment = center_align


# ── Sheet 1: Daily Plan ──

ws1 = wb.active
ws1.title = "Daily Plan"
ws1.sheet_properties.tabColor = DARK_BLUE

col_widths_1 = [5, 12, 6, 22, 70, 28, 14]
for i, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

apply_title_row(ws1, 1, "InnovateUS Public Engagement Coaching Tool — Daily Build Plan")
apply_subtitle_row(ws1, 2, "Feb 9 – Mar 16, 2026  |  Kushal + Manan + Batuhan  |  ~18 Working Days  |  OpenAI GPT-5.1 + Weaviate RAG")
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 22

apply_headers(ws1, 4, ["#", "Date", "Day", "Phase", "What We Did / Do", "Features", "Status"])

daily_data = [
    [1, "Feb 9", "Sun", "Setup & Research",
     "• Create Git repo & project structure\n• Research RAG pipeline options (Weaviate, OpenAI embeddings)\n• Study Rebooting Democracy codebase for adaptation\n• Set up .env config, .gitignore, LICENSE",
     "Planning + research", "Completed"],
    [2, "Feb 10", "Mon", "Ingestion Pipeline",
     "• Build basic public HTML chatbot UI\n• Create PDF ingestion pipeline (ingest-pdfs.mjs)\n• Set up Weaviate schema and connection\n• Implement LLM-based intelligent chunking strategy\n• Test chunking model variations",
     "#1 Ingestion pipeline\n#2 Basic chatbot UI", "Completed"],
    [3, "Feb 11", "Tue", "Ingestion Pipeline",
     "• Improve ingestion strategy V1 (chapter-based splitting)\n• Add PDF-to-Markdown conversion via Docling (convert-pdfs.py)\n• Implement BM25 + semantic vector search fallback\n• Refine Weaviate search queries",
     "#1 Ingestion V2", "Completed"],
    [4, "Feb 17", "Mon", "UI Build",
     "• Build full React app: Home, Coach, CaseStudies, CaseStudyDetail, ChatBot, Layout\n• Set up Vite + Tailwind CSS + TypeScript\n• Create Radix UI component library (Badge, Button, Card, Dialog, Tabs, etc.)\n• Add React Router with all routes\n• Build 9-question Nesta coaching assessment UI",
     "#3 React UI\n#5 Coach assessment\n#8 Case studies UI", "Completed"],
    [5, "Feb 18", "Tue", "Data Ingestion",
     "• Build Excel ingestion pipeline (ingest.py)\n• Ingest Participedia case studies from Data Tracker.xlsx\n• Implement section-based and sliding-window chunking for Excel data\n• Create CaseStudyLibrary collection in Weaviate",
     "#1 Excel pipeline\n#8 Case study data", "Completed"],
    [6, "Feb 19", "Wed", "Backend Integration",
     "• Wire UI to backend: chatbot, coach, case studies all connected\n• Implement RAG agent tools (search_knowledge_base, get_document_details)\n• Build agent runner with function calling loop (max 5 iterations)\n• Add case study data from Participedia (Batuhan)\n• Add Weaviate MCP for Cursor integration\n• Fix Netlify function timeouts\n• Adjust chunking/tagging strategy for better retrieval",
     "#2 RAG chatbot\n#4 Agent system\n#8 Case studies", "Completed"],
    [7, "Feb 20", "Thu", "Migration & Testing",
     "• Migrate from Netlify Functions to Express server (server.mjs)\n• Build all API endpoints: /api/chatbot, /api/adapt-case-study, /api/generate-plan (later pivoted)\n• Add relevancy score metric for case studies (Manan)\n• Add test suite: Vitest + React Testing Library (Manan)\n• Rebrand UI to InnovateUS theme\n• Create markdown-content component for rich rendering",
     "#11 Relevancy scoring\n#14 Test suite\n#13 InnovateUS branding", "Completed"],
    [8, "Feb 23", "Fri", "Refinement & Deploy",
     "• Add search bar and topic filtering for case studies (Batuhan)\n• Refine case study output formatting\n• Set up Render deployment config (render.yaml)\n• Fix Render build command\n• Update .gitignore",
     "#12 Search & filter\n#15 Deployment config", "Completed"],
    [9, "Mar 3", "Mon", "Coach Flow Redesign",
     "• Redesign coach into multi-step flow: Assessment → Dashboard → Coaching → Reflection\n• Build AssessmentDashboard with per-question results and status tracking\n• Build CoachingChatPanel for per-question AI coaching conversations\n• Build Reflection page with PDF export (jsPDF)\n• Add new routes (/coach/dashboard, /coach/reflection)\n• Add AI evaluation endpoint (/api/evaluate-assessment)",
     "#5 Assessment dashboard\n#6 Coaching chat\n#7 Reflection + PDF", "Completed"],
    [10, "Mar 4", "Tue", "Full Flow + Polish",
     "• Polish assessment dashboard: status indicators, progress tracking\n• Improve coaching chat UX: conversation flow, streaming display\n• End-to-end test of full coaching flow: assessment → evaluation → coaching → reflection\n• Fix edge cases in assessment responses",
     "UX polish", "Not Started"],
    [11, "Mar 5", "Wed", "Full Flow + Polish",
     "• Polish case study adaptation flow (/api/adapt-case-study)\n• Improve case study scoring algorithm\n• Refine coaching chat prompts and AI evaluation quality\n• Test case study recommendation accuracy against user context",
     "#9 Adaptation flow\n#6 Coaching chat polish", "Not Started"],
    [12, "Mar 6", "Thu", "Full Flow + Polish",
     "• Accessibility check: keyboard nav, screen reader, color contrast\n• Mobile responsiveness pass on all components\n• Performance optimization: API response times, lazy loading\n• Security pass: input sanitization, API key handling",
     "Accessibility + security", "Not Started"],
    [13, "Mar 9", "Mon", "Testing",
     "• Run 15-20 full end-to-end test sessions\n• Test full coaching flow: assessment → AI eval → coaching → reflection → PDF\n• Test chatbot with diverse questions\n• Test case study search, filter, and adaptation\n• Document all bugs: P0 / P1 / P2",
     "E2E testing", "Not Started"],
    [14, "Mar 10", "Tue", "Testing",
     "• Fix P0 and P1 bugs\n• AI accuracy evaluation: test coaching quality across question types\n• Test with simulated concurrent sessions\n• Expand test suite coverage for new components",
     "Bug fixes + AI eval", "Not Started"],
    [15, "Mar 11", "Wed", "Testing",
     "• Beta test with 3-5 InnovateUS users\n• Observe and collect feedback on coaching flow\n• Test case study relevancy and recommendations\n• Fix critical issues found during beta",
     "Beta testing", "Not Started"],
    [16, "Mar 12", "Thu", "Deploy",
     "• Implement critical beta feedback\n• Deploy to production (Render)\n• Run smoke tests on production URL\n• Final content review: all copy, questions, labels",
     "Deploy to production", "Not Started"],
    [17, "Mar 13", "Fri", "Deploy",
     "• Final QA pass on production\n• Build demo: screen recordings, walkthrough, key metrics\n• Create/update technical documentation & setup guide\n• Update README.md to reflect current architecture",
     "Final QA + demo prep", "Not Started"],
    [18, "Mar 16", "Mon", "Buffer",
     "• Emergency bug fixes\n• Practice demo presentation\n• Prepare handoff documentation\n• PROTOTYPE COMPLETE — ready for pilot",
     "Buffer / handoff\nPROTOTYPE DONE", "Not Started"],
]

for idx, row_data in enumerate(daily_data):
    r = 5 + idx
    apply_data_row(ws1, r, row_data, is_alt=(idx % 2 == 1))
    ws1.row_dimensions[r].height = 90
    apply_status_color(ws1, r, 7)

total_row = 5 + len(daily_data)
ws1.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
cell = ws1.cell(row=total_row, column=1, value="TOTAL")
cell.font = bold_font
cell.fill = light_fill
cell.border = thin_border
for c in range(2, 4):
    ws1.cell(row=total_row, column=c).fill = light_fill
    ws1.cell(row=total_row, column=c).border = thin_border
ws1.cell(row=total_row, column=4, value="18 days").font = bold_font
ws1.cell(row=total_row, column=4).fill = light_fill
ws1.cell(row=total_row, column=4).border = thin_border
ws1.cell(row=total_row, column=5, value="Strategy: OpenAI GPT-5.1 + Weaviate RAG from day one → Express backend → Render deploy").font = bold_font
ws1.cell(row=total_row, column=5).fill = light_fill
ws1.cell(row=total_row, column=5).border = thin_border
ws1.cell(row=total_row, column=5).alignment = wrap_align
for c in [6, 7]:
    ws1.cell(row=total_row, column=c).fill = light_fill
    ws1.cell(row=total_row, column=c).border = thin_border

# ── Sheet 2: Tech Stack ──

ws2 = wb.create_sheet("Tech Stack")
ws2.sheet_properties.tabColor = MED_BLUE

col_widths_2 = [22, 38, 50, 14, 45]
for i, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

apply_title_row(ws2, 1, "Tech Stack — Current Implementation", cols=5)
apply_subtitle_row(ws2, 2, "Strategy: Built on OpenAI + Weaviate RAG from day one. Migrated from Netlify to Express in Week 2.", cols=5)
ws2.row_dimensions[1].height = 30
ws2.row_dimensions[2].height = 22

apply_headers(ws2, 4, ["Category", "Using", "Why", "Status", "Alternatives Considered"])

tech_data = [
    ["Frontend", "React 18 + Vite 6 + Tailwind CSS 4 + TypeScript", "Fast build, type safety, rapid styling, modern tooling", "Confirmed", "Next.js, plain HTML"],
    ["UI Components", "Radix UI (Dialog, Tabs, ScrollArea, Select, Tooltip)", "Accessible, unstyled primitives, composable with Tailwind", "Confirmed", "shadcn/ui, MUI, Chakra UI"],
    ["Icons", "Lucide React", "Lightweight, consistent, tree-shakeable icon set", "Confirmed", "Heroicons, FontAwesome"],
    ["Markdown Rendering", "react-markdown + remark-gfm", "Rich content display for AI responses, GFM table support", "Confirmed", "MDX, marked"],
    ["PDF Export", "jsPDF", "Client-side PDF generation for reflection reports", "Confirmed", "html2pdf, Puppeteer"],
    ["Routing", "React Router 7", "Standard React routing with nested routes", "Confirmed", "TanStack Router"],
    ["Backend", "Express 5 (Node.js)", "Simple REST APIs, SSE streaming support, JS end-to-end", "Confirmed", "Netlify Functions (abandoned), FastAPI"],
    ["Vector Database", "Weaviate (Cloud)", "Hybrid search (BM25 + vector), built-in OpenAI vectorizer module", "Confirmed", "Pinecone, Qdrant, ChromaDB"],
    ["AI / LLM", "OpenAI GPT-5.1 / GPT-5.1-mini", "High quality coaching responses, function calling for agent tools", "Confirmed", "Claude, Gemini"],
    ["Embeddings", "text-embedding-3-small (via Weaviate)", "Cost-efficient, good quality, integrated with Weaviate vectorizer", "Confirmed", "text-embedding-3-large"],
    ["PDF Ingestion", "Docling (Python) + custom Node.js pipeline", "PDF → Markdown → LLM-based chunking → Weaviate", "Confirmed", "LangChain, LlamaIndex"],
    ["Excel Ingestion", "pandas + openpyxl (Python)", "Participedia case study data from Excel → Weaviate", "Confirmed", "SheetJS"],
    ["Testing", "Vitest + @testing-library/react + jsdom", "Fast Vite-native testing with component testing", "Confirmed", "Jest, Playwright"],
    ["Hosting", "Render", "Auto-deploy, supports Express server, free tier available", "Planned", "Vercel, Railway, Netlify"],
    ["Dev Tools", "Weaviate MCP (Cursor integration)", "Direct DB queries and schema inspection from IDE", "Confirmed", "Weaviate Console"],
]

for idx, row_data in enumerate(tech_data):
    r = 5 + idx
    apply_data_row(ws2, r, row_data, is_alt=(idx % 2 == 1))
    ws2.row_dimensions[r].height = 35
    apply_status_color(ws2, r, 4)

note_row = 5 + len(tech_data) + 1
ws2.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
cell = ws2.cell(row=note_row, column=1, value="KEY DECISION: Migrated from Netlify Functions to Express on Feb 20 for better SSE streaming, simpler deployment, and unified server.")
cell.font = Font(name="Calibri", size=10, bold=True, italic=True)
cell.fill = light_fill
cell.alignment = wrap_align
for c in range(2, 6):
    ws2.cell(row=note_row, column=c).fill = light_fill

# ── Sheet 3: Feature Tracker ──

ws3 = wb.create_sheet("Feature Tracker")
ws3.sheet_properties.tabColor = "2E8B57"

col_widths_3 = [5, 50, 22, 16, 14]
for i, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

apply_title_row(ws3, 1, "15 Core Features — Completion Tracker (includes 1 pivoted)", cols=5)
ws3.row_dimensions[1].height = 30

apply_headers(ws3, 3, ["#", "Feature", "Target Phase", "Target Done By", "Status"])

feature_data = [
    [1, "Knowledge base ingestion (PDF + Excel → Weaviate)", "Ingestion Pipeline", "Feb 18", "Completed"],
    [2, "RAG Chatbot with SSE streaming & source citations", "Backend Integration", "Feb 19", "Completed"],
    [3, "Full React UI (Home, Coach, Case Studies, Layout)", "UI Build", "Feb 17", "Completed"],
    [4, "Agent system (function calling, tool loop, hybrid search)", "Backend Integration", "Feb 19", "Completed"],
    [5, "Nesta 9-question coaching assessment + AI evaluation", "Coach Flow Redesign", "Mar 3", "Completed"],
    [6, "Per-question coaching chat panel", "Coach Flow Redesign", "Mar 3", "Completed"],
    [7, "Reflection page with PDF export", "Coach Flow Redesign", "Mar 3", "Completed"],
    [8, "Case study library (browse, filter, detail view)", "Backend Integration", "Feb 19", "Completed"],
    [9, "Case study AI adaptation (\"Adapt to my situation\")", "Backend Integration", "Feb 20", "Completed"],
    [10, "Engagement plan generator (questionnaire → follow-ups → plan) — PIVOTED to coaching flow", "Backend Integration", "Feb 20", "Pivoted"],
    [11, "Relevancy scoring for case studies", "Migration & Testing", "Feb 20", "Completed"],
    [12, "Search bar + topic filtering for case studies", "Refinement", "Feb 23", "Completed"],
    [13, "InnovateUS UI branding & theming", "Migration & Testing", "Feb 20", "Completed"],
    [14, "Test suite (component tests for all major views)", "Migration & Testing", "Feb 20", "Completed"],
    [15, "Deployment configuration (Render)", "Refinement", "Feb 23", "Started"],
]

for idx, row_data in enumerate(feature_data):
    r = 4 + idx
    apply_data_row(ws3, r, row_data, is_alt=(idx % 2 == 1))
    ws3.row_dimensions[r].height = 28
    apply_status_color(ws3, r, 5)

# ── Sheet 4: Phase Summary ──

ws4 = wb.create_sheet("Phase Summary")
ws4.sheet_properties.tabColor = "8B4513"

col_widths_4 = [24, 16, 8, 55, 45]
for i, w in enumerate(col_widths_4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

apply_title_row(ws4, 1, "Phase Summary", cols=5)
ws4.row_dimensions[1].height = 30

apply_headers(ws4, 3, ["Phase", "Dates", "Days", "What Got / Gets Done", "Milestone"])

phase_data = [
    ["Setup & Research", "Feb 9", "1",
     "Repo setup, architecture decisions, license, .gitignore, initial README",
     "Dev environment ready, repo initialized"],
    ["Ingestion Pipeline", "Feb 10–11", "2",
     "PDF ingestion, LLM-based chunking, Weaviate schema, Docling PDF→MD, basic chatbot UI",
     "Documents ingested into Weaviate, chatbot functional"],
    ["UI Build", "Feb 17", "1",
     "Full React app with all pages, Radix UI component library, routing, Nesta assessment, Tailwind styling",
     "Complete frontend scaffold with all views"],
    ["Data Ingestion & Backend", "Feb 18–19", "2",
     "Excel pipeline, case study data, RAG agent tools, agent runner, wired UI to backend, chunking adjustments",
     "Full-stack app: UI connected to RAG backend"],
    ["Migration & Testing", "Feb 20", "1",
     "Netlify → Express migration, test suite, relevancy scoring, InnovateUS branding, markdown rendering",
     "Express server live, tests passing, branded UI"],
    ["Refinement & Deploy Config", "Feb 23", "1",
     "Search & filter for case studies, Render.yaml, output formatting",
     "Case study discovery improved, deploy config ready"],
    ["Coach Flow Redesign", "Mar 3", "1",
     "Multi-step coaching: Assessment → Dashboard → Coaching Chat → Reflection, PDF export, new routes",
     "Complete AI coaching pipeline working"],
    ["Full Flow + Polish", "Mar 4–6", "3",
     "End-to-end integration, UX polish, accessibility, security, performance, coaching flow & case study refinement",
     "Participant flow polished & secure"],
    ["Testing + Deploy", "Mar 9–13", "5",
     "E2E testing, bug fixes, beta with InnovateUS users, production deploy, demo preparation",
     "All 15 features verified in production"],
    ["Buffer", "Mar 16", "1",
     "Emergency fixes, demo practice, handoff docs",
     "PROTOTYPE COMPLETE"],
]

for idx, row_data in enumerate(phase_data):
    r = 4 + idx
    apply_data_row(ws4, r, row_data, is_alt=(idx % 2 == 1))
    ws4.row_dimensions[r].height = 45

# ── Sheet 5: Deliverables ──

ws5 = wb.create_sheet("Deliverables")
ws5.sheet_properties.tabColor = "8B008B"

col_widths_5 = [5, 30, 70, 14, 16, 20, 14]
for i, w in enumerate(col_widths_5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

apply_title_row(ws5, 1, "Summary of Deliverables — For Review by Beth & Advisors")
ws5.row_dimensions[1].height = 30

apply_headers(ws5, 3, ["#", "Deliverable", "Description / What to Review", "Ready By", "Type", "Reviewer", "Status"])

deliverable_data = [
    [1, "Project Plan & Timeline",
     "Daily plan, tech stack, feature list, phase breakdown. Confirms scope, timeline, and division of work across Kushal, Manan, and Batuhan.",
     "Feb 9", "Document", "Beth, Ani, David", "Completed"],
    [2, "Knowledge Base Pipeline",
     "Working PDF + Excel ingestion into Weaviate. LLM-based intelligent chunking with review loop, hybrid BM25 + vector search. Participedia case studies loaded.",
     "Feb 18", "Technical Demo", "Beth, Ani, David", "Completed"],
    [3, "RAG Chatbot Demo",
     "Floating Q&A chatbot: SSE streaming, source citations, contextual answers from knowledge base. Demonstrates retrieval quality.",
     "Feb 19", "Demo / Gate", "Beth, Ani, David", "Completed"],
    [4, "Full UI Prototype",
     "React app with all pages: Home, Coach (9 Nesta questions), Case Studies (browse/filter/detail), ChatBot. InnovateUS branding applied.",
     "Feb 20", "Demo / Gate", "Beth, Ani, David", "Completed"],
    [5, "Express Migration & Test Suite",
     "Backend migrated from Netlify Functions to Express. All API endpoints working. Component tests for Home, Layout, Coach, ChatBot, CaseStudies.",
     "Feb 20", "Technical Milestone", "Beth, Ani, David", "Completed"],
    [6, "Case Study Library with Search",
     "Case studies from Weaviate with topic filtering, search bar, relevancy scoring, and AI adaptation flow (\"Adapt to my situation\").",
     "Feb 23", "Demo / Gate", "Beth, Ani, David", "Completed"],
    [7, "AI Coaching Flow",
     "Complete multi-step flow: 9-question Nesta assessment → AI evaluation (addressed/partial/not-addressed) → per-question coaching chat → reflection with PDF export.",
     "Mar 3", "Demo / Gate", "Beth, Ani, David", "Completed"],
    [8, "Engagement Plan Generator (Pivoted)",
     "Originally built as questionnaire → follow-up questions → structured plan. Pivoted to the Nesta coaching flow which provides more structured, per-question AI coaching with assessment evaluation and reflection.",
     "Feb 20", "Pivoted Feature", "Beth, Ani, David", "Pivoted"],
    [9, "Polished Participant Experience",
     "Accessibility (WCAG 2.1 AA), mobile responsiveness, security (input sanitization, API key handling), performance optimization.",
     "Mar 6", "Review / Gate", "Beth, Ani, David", "Not Started"],
    [10, "E2E Test Results",
     "Summary of 15-20 end-to-end test sessions across devices. Bug documentation (P0/P1/P2) and resolution status.",
     "Mar 10", "Report", "Beth, Ani, David", "Not Started"],
    [11, "Beta Test Results",
     "Feedback from 3-5 InnovateUS beta users: usability, coaching quality, completion flow, pain points, improvements made.",
     "Mar 11", "Report", "Beth, Ani, David", "Not Started"],
    [12, "Production Deployment",
     "Prototype live on Render production URL. All 15 features verified. Smoke test results.",
     "Mar 12", "Milestone", "Beth, Ani, David", "Not Started"],
    [13, "Demo Package",
     "Live walkthrough script, screen recordings of coaching flow + case studies + chatbot, architecture diagram, AI quality metrics.",
     "Mar 13", "Presentation", "Beth, Ani, David", "Not Started"],
    [14, "Technical Documentation",
     "Updated README, API docs, architecture overview, deployment instructions, environment config guide. Everything needed to maintain the system.",
     "Mar 13", "Document", "Beth, Ani, David", "Not Started"],
    [15, "Final Prototype Handoff",
     "Working production URL, source code repo, documentation, test results, demo materials. Ready for pilot with InnovateUS cohort.",
     "Mar 16", "Final Handoff", "Beth, Ani, David", "Not Started"],
]

for idx, row_data in enumerate(deliverable_data):
    r = 4 + idx
    apply_data_row(ws5, r, row_data, is_alt=(idx % 2 == 1))
    ws5.row_dimensions[r].height = 55
    apply_status_color(ws5, r, 7)

note_row2 = 4 + len(deliverable_data) + 1
ws5.merge_cells(start_row=note_row2, start_column=1, end_row=note_row2, end_column=7)
cell = ws5.cell(row=note_row2, column=1,
                value="GATES = Checkpoints where advisor go/no-go feedback is needed before proceeding to the next phase")
cell.font = Font(name="Calibri", size=10, bold=True, italic=True)
cell.fill = light_fill
cell.alignment = wrap_align
for c in range(2, 8):
    ws5.cell(row=note_row2, column=c).fill = light_fill

# ── Freeze panes ──
ws1.freeze_panes = "A5"
ws2.freeze_panes = "A5"
ws3.freeze_panes = "A4"
ws4.freeze_panes = "A4"
ws5.freeze_panes = "A4"

output_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                           "CoachingTool_Prototype_Plan.xlsx")
wb.save(output_path)
print(f"Saved to: {output_path}")
